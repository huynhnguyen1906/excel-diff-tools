"""
Haritsuke Sheet Excel Reader
貼付シート専用の Excel ファイル読み込みモジュール
"""
from pathlib import Path
from typing import Optional, Tuple, List
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime, timedelta


class HaritsukeExcelReader:
    """貼付シート用 Excel ファイル読み込みクラス"""
    
    def __init__(self, file_path: Path):
        """
        初期化
        
        Args:
            file_path: Excel ファイルのパス
        """
        self.file_path = file_path
        self._workbook = None
        self._sheet_names = []
    
    def validate_file(self) -> Tuple[bool, str]:
        """
        ファイルが有効な Excel ファイルかどうかを検証
        
        Returns:
            (is_valid, error_message): 
                - is_valid: True なら有効
                - error_message: エラーがあればメッセージ、なければ空文字列
        """
        # ファイルの存在確認
        if not self.file_path.exists():
            return False, f"ファイルが見つかりません: {self.file_path.name}"
        
        # ファイル拡張子の確認
        if self.file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
            return False, f"対応していないファイル形式です: {self.file_path.suffix}\n(.xlsx または .xlsm のみ対応)"
        
        # ファイルを開いて検証
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
            
            if not self._sheet_names:
                return False, "シートが見つかりません"
            
            return True, ""
            
        except InvalidFileException:
            return False, "無効な Excel ファイルです"
        except PermissionError:
            return False, "ファイルを開けません（他のプログラムで使用中の可能性があります）"
        except Exception as e:
            return False, f"ファイル読み込みエラー: {str(e)}"
    
    def get_sheet_names(self) -> List[str]:
        """
        ファイル内のすべてのシート名を取得
        
        Returns:
            シート名のリスト
        """
        if not self._sheet_names:
            # まだ読み込んでいない場合は検証を実行
            is_valid, _ = self.validate_file()
            if not is_valid:
                return []
        
        return self._sheet_names
    
    def validate_sheet(self, sheet_name: str) -> Tuple[bool, str]:
        """
        指定されたシートが存在するか検証
        
        Args:
            sheet_name: 確認するシート名
            
        Returns:
            (exists, error_message)
        """
        if not sheet_name or sheet_name.strip() == "":
            return False, "シート名が空です"
        
        # シート名の大文字小文字を区別しない比較
        sheet_names_lower = [s.lower() for s in self.get_sheet_names()]
        
        if sheet_name.lower() not in sheet_names_lower:
            available = ', '.join(self._sheet_names[:5])  # 最初の5つまで表示
            if len(self._sheet_names) > 5:
                available += f" ... (他 {len(self._sheet_names) - 5} シート)"
            return False, f"シート '{sheet_name}' が見つかりません\n\n利用可能なシート:\n{available}"
        
        return True, ""
    
    def read_sheet(self, sheet_name: str) -> Tuple[Optional[pd.DataFrame], str]:
        """
        指定されたシートのデータを pandas DataFrame として読み込む
        
        Args:
            sheet_name: 読み込むシート名
            
        Returns:
            (dataframe, error_message):
                - dataframe: 成功時は DataFrame、失敗時は None
                - error_message: エラーがあればメッセージ
        """
        # シートの存在確認
        is_valid, error_msg = self.validate_sheet(sheet_name)
        if not is_valid:
            return None, error_msg
        
        try:
            # 実際のシート名を取得（大文字小文字の違いを吸収）
            actual_sheet_name = None
            for name in self._sheet_names:
                if name.lower() == sheet_name.lower():
                    actual_sheet_name = name
                    break
            
            # pandas で読み込み
            # header=0: 1行目をヘッダーとして使用
            df = pd.read_excel(
                self.file_path,
                sheet_name=actual_sheet_name,
                header=0,
                na_values=['', 'NA', 'N/A', 'null', 'NULL', 'None']
            )
            
            # datetime 列を column name で検出
            excel_epoch = datetime(1899, 12, 30)
            
            for col_name in df.columns:
                # 列名に "日時" が含まれる場合は datetime として扱う
                if '日時' in str(col_name):
                    # まだ datetime になっていない float 値を変換
                    df[col_name] = df[col_name].apply(lambda x: 
                        excel_epoch + timedelta(days=float(x)) 
                        if isinstance(x, (int, float)) and not pd.isna(x) and not isinstance(x, (pd.Timestamp, datetime))
                        else x
                    )
            
            # 空のDataFrameチェック
            if df.empty:
                return None, f"シート '{sheet_name}' にデータがありません"
            
            # 行数チェック（ヘッダー除く）
            if len(df) == 0:
                return None, f"シート '{sheet_name}' にデータ行がありません（ヘッダーのみ）"
            
            return df, ""
            
        except Exception as e:
            return None, f"シート読み込みエラー: {str(e)}"
    
    def get_sheet_info(self, sheet_name: str) -> Tuple[Optional[dict], str]:
        """
        シートの基本情報を取得（行数、列数など）
        
        Args:
            sheet_name: シート名
            
        Returns:
            (info_dict, error_message)
        """
        df, error_msg = self.read_sheet(sheet_name)
        if df is None:
            return None, error_msg
        
        info = {
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'has_data': not df.empty
        }
        
        return info, ""
