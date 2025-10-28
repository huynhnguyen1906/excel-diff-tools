"""
Haritsuke Excel Writer
貼付シート専用の Excel 書き込みモジュール - 差分結果を Excel ファイルとして出力
"""
from pathlib import Path
from datetime import datetime
from typing import List
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .diff_engine import HaritsukeDiffResult
from utils.constants import DIFF_COLORS


class HaritsukeExcelWriter:
    """貼付シート専用 - 差分結果を Excel ファイルとして出力するクラス"""
    
    def __init__(self, output_dir: Path, sheet_name: str):
        """
        初期化
        
        Args:
            output_dir: 出力ディレクトリ
            sheet_name: シート名
        """
        self.output_dir = output_dir
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        
        # カラー定義
        self.color_deleted = PatternFill(start_color=DIFF_COLORS['deleted'], 
                                        end_color=DIFF_COLORS['deleted'], 
                                        fill_type='solid')
        self.color_changed = PatternFill(start_color=DIFF_COLORS['changed'], 
                                        end_color=DIFF_COLORS['changed'], 
                                        fill_type='solid')
        self.color_added = PatternFill(start_color=DIFF_COLORS['added'], 
                                      end_color=DIFF_COLORS['added'], 
                                      fill_type='solid')
        
        # フォント定義
        self.font_normal = Font(name='Segoe UI', size=10)
        self.font_bold = Font(name='Segoe UI', size=10, bold=True)
        
        # ボーダー定義
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, 
                            top=thin_border, bottom=thin_border)
    
    def write_diff_results(self, 
                          columns: List[str], 
                          results: List[HaritsukeDiffResult]) -> Path:
        """
        差分結果を Excel ファイルに書き込み
        
        Args:
            columns: 列名のリスト
            results: 差分結果のリスト
            
        Returns:
            出力ファイルのパス
        """
        # Workbook 作成
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name
        
        # ヘッダー行を書き込み
        self._write_header(columns)
        
        # データ行を書き込み
        for result in results:
            self._write_diff_row(result, columns)
        
        # 書式設定
        self._apply_formatting(len(columns))
        
        # ファイル保存
        output_path = self._generate_output_path()
        self.wb.save(output_path)
        
        return output_path
    
    def _write_header(self, columns: List[str]):
        """
        ヘッダー行を書き込み
        
        Args:
            columns: 列名のリスト
        """
        for col_idx, col_name in enumerate(columns, start=1):
            cell = self.ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = self.font_bold
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _write_diff_row(self, result: HaritsukeDiffResult, columns: List[str]):
        """
        1行の差分データを書き込み
        
        Args:
            result: 差分結果
            columns: 列名のリスト
        """
        # 出力行番号 (+1 for header)
        row_num = result.row_index + 1
        
        # データソースを決定
        if result.change_type == 'deleted':
            data = result.old_data
        elif result.change_type == 'added':
            data = result.new_data
        else:  # changed
            data = result.new_data  # 新データを基準
        
        # 各セルを書き込み
        for col_idx, col_name in enumerate(columns, start=1):
            cell = self.ws.cell(row=row_num, column=col_idx)
            cell.font = self.font_normal
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = self.border  # ボーダーを追加
            
            # 変更タイプに応じた処理
            if result.change_type == 'deleted':
                    # 削除行: 全セル赤背景
                cell.fill = self.color_deleted
                value = data.get(col_name, '')
                # datetime を保持
                if isinstance(value, (pd.Timestamp, datetime)):
                    cell.value = value
                    cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                else:
                    cell.value = '' if pd.isna(value) else str(value)
                
            elif result.change_type == 'added':
                # 追加行: 全セル緑背景
                cell.fill = self.color_added
                value = data.get(col_name, '')
                # datetime を保持
                if isinstance(value, (pd.Timestamp, datetime)):
                    cell.value = value
                    cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                else:
                    cell.value = '' if pd.isna(value) else str(value)
                
            elif result.change_type == 'changed':
                # 変更行: 変更されたセルのみ色付け
                if col_name in result.changed_columns:
                    old_val = result.old_data.get(col_name, '')
                    new_val = result.new_data.get(col_name, '')
                    
                    # datetime チェック
                    is_datetime = isinstance(new_val, (pd.Timestamp, datetime)) or isinstance(old_val, (pd.Timestamp, datetime))
                    
                    if is_datetime:
                        # datetime の場合は rich text でなく直接値を設定
                        cell.fill = self.color_changed
                        if isinstance(new_val, (pd.Timestamp, datetime)):
                            cell.value = new_val
                            cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                        else:
                            cell.value = str(new_val) if not pd.isna(new_val) else ''
                    else:
                        # 旧値・新値の空チェック
                        old_is_empty = pd.isna(old_val) or str(old_val).strip() == ''
                        new_is_empty = pd.isna(new_val) or str(new_val).strip() == ''
                        
                        if old_is_empty and not new_is_empty:
                            # 旧: 空 → 新: 値あり → 緑背景（追加）
                            cell.fill = self.color_added
                            cell.value = str(new_val)
                        elif not old_is_empty and new_is_empty:
                            # 旧: 値あり → 新: 空 → 赤背景（削除）
                            cell.fill = self.color_deleted
                            cell.value = f"~{old_val}~"
                        else:
                            # 旧: 値あり → 新: 値あり（異なる）→ 黄背景（変更）
                            cell.fill = self.color_changed
                            cell.value = self._create_rich_text(old_val, new_val)
                else:
                    # 変更されていないセル: 背景色なし
                    value = data.get(col_name, '')
                    # datetime を保持
                    if isinstance(value, (pd.Timestamp, datetime)):
                        cell.value = value
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                    else:
                        cell.value = '' if pd.isna(value) else str(value)
    
    def _create_rich_text(self, old_value, new_value):
        """
        セルのリッチテキストを作成（~old~ → new の形式）
        
        Note: openpyxl の rich text API に制限があるため、
        "~旧値~ → 新値" という文字列表現を使用
        
        Args:
            old_value: 旧値
            new_value: 新値
            
        Returns:
            文字列（"~旧値~ → 新値" の形式）
        """
        # 値を文字列に変換
        old_str = '' if pd.isna(old_value) else str(old_value)
        new_str = '' if pd.isna(new_value) else str(new_value)
        
        # ~old~ → new の形式で返す
        # 取り消し線は ~ で表現
        if old_str and new_str:
            return f"~{old_str}~ → {new_str}"
        elif old_str:
            return f"~{old_str}~"
        elif new_str:
            return new_str
        else:
            return ""
    
    def _apply_formatting(self, num_columns: int):
        """
        ワークシート全体の書式を適用
        
        Args:
            num_columns: 列数
        """
        # Freeze panes (1行目を固定)
        self.ws.freeze_panes = 'A2'
        
        # AutoFilter (ヘッダー行)
        self.ws.auto_filter.ref = f'A1:{get_column_letter(num_columns)}1'
        
        # 列幅の自動調整
        self._auto_fit_columns(num_columns)
    
    def _auto_fit_columns(self, num_columns: int):
        """
        列幅を自動調整 (最大100文字)
        
        Args:
            num_columns: 列数
        """
        MAX_WIDTH = 100
        
        for col_idx in range(1, num_columns + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            
            # 列内の最大文字数を計算
            for row_idx in range(1, self.ws.max_row + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                
                # セルの値の文字数を計算
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # 列幅を設定 (最小10, 最大100)
            adjusted_width = min(max(max_length + 2, 10), MAX_WIDTH)
            self.ws.column_dimensions[col_letter].width = adjusted_width
    
    def _generate_output_path(self) -> Path:
        """
        出力ファイルパスを生成
        
        Returns:
            出力ファイルのパス
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f"diff_{self.sheet_name}_{timestamp}.xlsx"
        return self.output_dir / filename
