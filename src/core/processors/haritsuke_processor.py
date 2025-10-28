"""
Haritsuke (貼付) Sheet Processor
貼付 シートの差分処理を担当するモジュール
"""
from pathlib import Path
from typing import Optional, Tuple, List, Callable
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime, timedelta

from .base_processor import BaseProcessor
from ..data_normalizer import DataNormalizer
from ..diff_engine import DiffEngine
from ..excel_writer import ExcelWriter


class HaritsukeExcelReader:
    """貼付シート用 Excel ファイル読み込みクラス"""
    
    def __init__(self, file_path: Path):
        """
        初期化
        
        Args:
            file_path: Excel ファイルのパス
        """
        self.file_path = file_path
        self._workbook = None
        self._sheet_names = []
    
    def validate_file(self) -> Tuple[bool, str]:
        """
        ファイルが有効な Excel ファイルかどうかを検証
        
        Returns:
            (is_valid, error_message): 
                - is_valid: True なら有効
                - error_message: エラーがあればメッセージ、なければ空文字列
        """
        # ファイルの存在確認
        if not self.file_path.exists():
            return False, f"ファイルが見つかりません: {self.file_path.name}"
        
        # ファイル拡張子の確認
        if self.file_path.suffix.lower() not in ['.xlsx', '.xlsm']:
            return False, f"対応していないファイル形式です: {self.file_path.suffix}\n(.xlsx または .xlsm のみ対応)"
        
        # ファイルを開いて検証
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            self._sheet_names = wb.sheetnames
            wb.close()
            
            if not self._sheet_names:
                return False, "シートが見つかりません"
            
            return True, ""
            
        except InvalidFileException:
            return False, "無効な Excel ファイルです"
        except PermissionError:
            return False, "ファイルを開けません（他のプログラムで使用中の可能性があります）"
        except Exception as e:
            return False, f"ファイル読み込みエラー: {str(e)}"
    
    def get_sheet_names(self) -> List[str]:
        """
        ファイル内のすべてのシート名を取得
        
        Returns:
            シート名のリスト
        """
        if not self._sheet_names:
            # まだ読み込んでいない場合は検証を実行
            is_valid, _ = self.validate_file()
            if not is_valid:
                return []
        
        return self._sheet_names
    
    def validate_sheet(self, sheet_name: str) -> Tuple[bool, str]:
        """
        指定されたシートが存在するか検証
        
        Args:
            sheet_name: 確認するシート名
            
        Returns:
            (exists, error_message)
        """
        if not sheet_name or sheet_name.strip() == "":
            return False, "シート名が空です"
        
        # シート名の大文字小文字を区別しない比較
        sheet_names_lower = [s.lower() for s in self.get_sheet_names()]
        
        if sheet_name.lower() not in sheet_names_lower:
            available = ', '.join(self._sheet_names[:5])  # 最初の5つまで表示
            if len(self._sheet_names) > 5:
                available += f" ... (他 {len(self._sheet_names) - 5} シート)"
            return False, f"シート '{sheet_name}' が見つかりません\n\n利用可能なシート:\n{available}"
        
        return True, ""
    
    def read_sheet(self, sheet_name: str) -> Tuple[Optional[pd.DataFrame], str]:
        """
        指定されたシートのデータを pandas DataFrame として読み込む
        
        Args:
            sheet_name: 読み込むシート名
            
        Returns:
            (dataframe, error_message):
                - dataframe: 成功時は DataFrame、失敗時は None
                - error_message: エラーがあればメッセージ
        """
        # シートの存在確認
        is_valid, error_msg = self.validate_sheet(sheet_name)
        if not is_valid:
            return None, error_msg
        
        try:
            # 実際のシート名を取得（大文字小文字の違いを吸収）
            actual_sheet_name = None
            for name in self._sheet_names:
                if name.lower() == sheet_name.lower():
                    actual_sheet_name = name
                    break
            
            # pandas で読み込み
            # header=0: 1行目をヘッダーとして使用
            df = pd.read_excel(
                self.file_path,
                sheet_name=actual_sheet_name,
                header=0,
                na_values=['', 'NA', 'N/A', 'null', 'NULL', 'None']
            )
            
            # datetime 列を column name で検出
            excel_epoch = datetime(1899, 12, 30)
            
            for col_name in df.columns:
                # 列名に "日時" が含まれる場合は datetime として扱う
                if '日時' in str(col_name):
                    # まだ datetime になっていない float 値を変換
                    df[col_name] = df[col_name].apply(lambda x: 
                        excel_epoch + timedelta(days=float(x)) 
                        if isinstance(x, (int, float)) and not pd.isna(x) and not isinstance(x, (pd.Timestamp, datetime))
                        else x
                    )
            
            # 空のDataFrameチェック
            if df.empty:
                return None, f"シート '{sheet_name}' にデータがありません"
            
            # 行数チェック（ヘッダー除く）
            if len(df) == 0:
                return None, f"シート '{sheet_name}' にデータ行がありません（ヘッダーのみ）"
            
            return df, ""
            
        except Exception as e:
            return None, f"シート読み込みエラー: {str(e)}"
    
    def get_sheet_info(self, sheet_name: str) -> Tuple[Optional[dict], str]:
        """
        シートの基本情報を取得（行数、列数など）
        
        Args:
            sheet_name: シート名
            
        Returns:
            (info_dict, error_message)
        """
        df, error_msg = self.read_sheet(sheet_name)
        if df is None:
            return None, error_msg
        
        info = {
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': df.columns.tolist(),
            'has_data': not df.empty
        }
        
        return info, ""


class HaritsukeProcessor(BaseProcessor):
    """貼付シート専用プロセッサ"""
    
    SHEET_NAME = "貼付"
    RECORD_NUMBER_COLUMN = 1  # Column B (0-indexed)
    
    def get_sheet_name(self) -> str:
        """対応するシート名を返す"""
        return self.SHEET_NAME
    
    def process(
        self, 
        old_file: Path, 
        new_file: Path, 
        sheet_name: str, 
        output_dir: Path,
        progress_callback: Optional[Callable[[int, str], None]] = None
    ) -> Tuple[Optional[Path], Optional[List], str]:
        """
        貼付シートの差分処理を実行
        
        Args:
            old_file: 旧ファイルのパス
            new_file: 新ファイルのパス
            sheet_name: シート名
            output_dir: 出力先ディレクトリ
            progress_callback: 進捗コールバック関数 (progress%, message)
            
        Returns:
            (output_path, diff_results, error_message):
                - output_path: 成功時は出力ファイルパス、失敗時は None
                - diff_results: 差分結果のリスト、失敗時は None
                - error_message: エラーがあればメッセージ
        """
        try:
            # Step 1: ファイル読み込み
            if progress_callback:
                progress_callback(20, "ファイルを読み込んでいます...")
            
            old_reader = HaritsukeExcelReader(old_file)
            new_reader = HaritsukeExcelReader(new_file)
            
            old_df, error_msg = old_reader.read_sheet(sheet_name)
            if old_df is None:
                return None, None, f"旧ファイル読み込みエラー:\n{error_msg}"
            
            new_df, error_msg = new_reader.read_sheet(sheet_name)
            if new_df is None:
                return None, None, f"新ファイル読み込みエラー:\n{error_msg}"
            
            # Step 2: データ正規化
            if progress_callback:
                progress_callback(40, "データを正規化しています...")
            
            normalizer = DataNormalizer()
            old_df_normalized = normalizer.normalize_dataframe(old_df)
            new_df_normalized = normalizer.normalize_dataframe(new_df)
            
            # 列を揃える
            old_df_aligned, new_df_aligned = normalizer.align_columns(
                old_df_normalized, new_df_normalized
            )
            
            # Step 3: 差分検出
            if progress_callback:
                progress_callback(60, "差分を検出しています...")
            
            diff_results = DiffEngine.compare_dataframes(
                old_df_aligned, new_df_aligned
            )
            
            # Step 4: Excel 出力
            if progress_callback:
                progress_callback(80, "結果を出力しています...")
            
            writer = ExcelWriter(output_dir, sheet_name)
            columns = new_df_aligned.columns.tolist()
            output_path = writer.write_diff_results(columns, diff_results)
            
            if progress_callback:
                progress_callback(100, "完了しました")
            
            return output_path, diff_results, ""
            
        except Exception as e:
            return None, None, f"処理中にエラーが発生しました:\n{str(e)}"
