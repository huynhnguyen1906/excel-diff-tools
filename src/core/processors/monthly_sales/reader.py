"""
Monthly Sales Excel Reader
月別売上２シート専用の Excel 読み取りモジュール
"""
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class MonthlySalesExcelReader:
    """月別売上２シート専用 - Excel ファイルを読み取るクラス"""
    
    # 定数
    HEADER_ROWS = 6  # ヘッダーは6行
    CATEGORY_COLUMNS = 3  # カテゴリ列はA, B, C (3列)
    DATA_START_ROW = 7  # データは7行目から開始
    DATA_START_COLUMN = 3  # データは D列 (0-indexed: 3) から開始
    BLOCK_SIZE = 4  # 1ヶ月のブロックは4列
    MONTH_ROW = 5  # 月名は5行目 (1-indexed)
    
    def __init__(self):
        """初期化"""
        pass
    
    def validate_file(self, file_path: Path) -> Tuple[bool, Optional[str]]:
        """
        ファイルの妥当性をチェック
        
        Args:
            file_path: ファイルパス
            
        Returns:
            (成功フラグ, エラーメッセージ)
        """
        if not file_path.exists():
            return False, f"ファイルが見つかりません: {file_path}"
        
        if not file_path.suffix.lower() in ['.xlsx', '.xlsm']:
            return False, f"サポートされていないファイル形式です: {file_path.suffix}"
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            wb.close()
            return True, None
        except Exception as e:
            return False, f"ファイルを開けませんでした: {str(e)}"
    
    def get_sheet_names(self, file_path: Path) -> List[str]:
        """
        ファイル内のシート名リストを取得
        
        Args:
            file_path: ファイルパス
            
        Returns:
            シート名のリスト
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    
    def validate_sheet(self, file_path: Path, sheet_name: str) -> Tuple[bool, Optional[str]]:
        """
        シートの妥当性をチェック
        
        Args:
            file_path: ファイルパス
            sheet_name: シート名
            
        Returns:
            (成功フラグ, エラーメッセージ)
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return False, f"シート '{sheet_name}' が見つかりません"
            
            wb.close()
            return True, None
            
        except Exception as e:
            print(f"[ERROR] validate_sheet エラー: {str(e)}")
            return False, f"シート検証エラー: {str(e)}"
    
    def _extract_month_blocks(self, ws: Worksheet, max_col: int = 50) -> Dict[str, Tuple[int, int]]:
        """
        月ブロックの情報を抽出 (月名 → (開始列index, 終了列index))
        
        Args:
            ws: ワークシート
            max_col: スキャンする最大列数 (デフォルト50列)
            
        Returns:
            {月名: (start_col, end_col)} の辞書
        """
        month_blocks = {}
        
        # 5行目 (MONTH_ROW) をスキャンして月名を探す
        current_col = self.DATA_START_COLUMN  # D列 (index 3) から開始
        
        # 最大列数を制限してスキャン（無限ループ防止）
        while current_col < max_col:
            # openpyxl は 1-indexed なので +1
            cell_value = ws.cell(row=self.MONTH_ROW, column=current_col + 1).value
            
            if cell_value and isinstance(cell_value, str):
                # "2024/12月" のような形式を期待
                if '月' in cell_value or '/' in cell_value:
                    # ブロックの終了列 = 開始列 + 3 (4列分)
                    end_col = current_col + self.BLOCK_SIZE - 1
                    month_blocks[cell_value] = (current_col, end_col)
                    current_col += self.BLOCK_SIZE
                else:
                    current_col += 1
            elif cell_value is None:
                # 空白セルが続いたら終了
                current_col += 1
                # 5列連続空白なら終了
                empty_count = 1
                while empty_count < 5 and current_col < max_col:
                    if ws.cell(row=self.MONTH_ROW, column=current_col + 1).value is None:
                        empty_count += 1
                        current_col += 1
                    else:
                        break
                if empty_count >= 5:
                    break
            else:
                current_col += 1
        
        return month_blocks
    
    def read_sheet(self, file_path: Path, sheet_name: str) -> Dict:
        """
        シートを読み取り、構造化されたデータを返す
        
        Args:
            file_path: ファイルパス
            sheet_name: シート名
            
        Returns:
            {
                'header': DataFrame (行1-6),
                'categories': DataFrame (列A-C, 全行),
                'month_blocks': {月名: DataFrame},
                'month_order': [月名リスト]
            }
        """
        try:
            
            # pandas で全データを読み込み（header なしで生データとして読む）
            df_raw = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=None,  # ヘッダーなし（全て生データ）
                na_values=['', 'NA', 'N/A', 'null', 'NULL', 'None']
            )
            
            
            # ヘッダー部分を抽出 (行0-5, 0-indexed)
            header_df = df_raw.iloc[0:self.HEADER_ROWS, :].copy()
            
            # カテゴリ列を抽出 (列0-2, 全行)
            category_df = df_raw.iloc[:, 0:self.CATEGORY_COLUMNS].copy()
            category_df.columns = ['A', 'B', 'C']
            
            # 月ブロックを抽出（5行目＝index 4から月名を探す）
            month_row = df_raw.iloc[self.MONTH_ROW - 1, :]  # 5行目 = index 4
            month_blocks = {}
            month_order = []
            
            current_col = self.DATA_START_COLUMN  # D列 = index 3
            
            for col_idx in range(current_col, len(month_row)):
                cell_value = month_row.iloc[col_idx]
                
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    # "2024/12月" のような月形式を検出
                    if '月' in cell_value and '/' in cell_value:
                        month_name = cell_value
                        start_col = col_idx
                        end_col = col_idx + self.BLOCK_SIZE - 1
                        month_blocks[month_name] = (start_col, end_col)
                        month_order.append(month_name)
                    # "全体の売上" のような全体ブロックを検出
                    elif cell_value.startswith('全体の'):
                        # 全体ブロックの最初の列（売上）を検出したら
                        if '売上' in cell_value:
                            month_name = '全体'  # 全体ブロックの名前
                            start_col = col_idx
                            end_col = col_idx + self.BLOCK_SIZE - 1
                            month_blocks[month_name] = (start_col, end_col)
                            month_order.append(month_name)
            
            
            if not month_blocks:
                raise ValueError(f"月ブロックが見つかりませんでした。5行目に '月' を含む列が必要です。")
            
            # 各月ブロックのデータを抽出
            month_data = {}
            
            for month_name, (start_col, end_col) in month_blocks.items():
                # データ行 (7行目以降 = index 6以降) を抽出
                block_df = df_raw.iloc[self.DATA_START_ROW - 1:, start_col:end_col + 1].copy()
                
                # 列名を設定
                block_df.columns = ['売上', '外部原価', '内部原価', '営業利益']
                
                # 数値に変換
                for col in block_df.columns:
                    block_df[col] = pd.to_numeric(block_df[col], errors='coerce')
                
                # indexをリセット
                block_df.reset_index(drop=True, inplace=True)
                
                month_data[month_name] = block_df
            
            
            return {
                'header': header_df,
                'categories': category_df,
                'month_blocks': month_data,
                'month_order': month_order
            }
        
        except Exception as e:
            print(f"[ERROR] read_sheet() でエラー発生: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"シート読み込みエラー: {str(e)}") from e
    
    def get_sheet_info(self, file_path: Path, sheet_name: str) -> Dict:
        """
        シート情報を取得 (デバッグ用)
        
        Args:
            file_path: ファイルパス
            sheet_name: シート名
            
        Returns:
            シート情報の辞書
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        info = {
            'total_rows': ws.max_row,
            'total_columns': ws.max_column,
            'data_rows': ws.max_row - self.HEADER_ROWS,
            'month_blocks': self._extract_month_blocks(ws)
        }
        
        wb.close()
        return info
