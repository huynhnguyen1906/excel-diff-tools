"""
Monthly Sales Excel Writer
月別売上２シート専用の Excel 書き込みモジュール - 差分結果を Excel ファイルとして出力
"""
from pathlib import Path
from typing import List, Dict
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .diff_engine import MonthlySalesMonthDiff, MonthlySalesCellDiff
from utils.constants import DIFF_COLORS


class MonthlySalesExcelWriter:
    """月別売上２シート専用 - 差分結果を Excel ファイルとして出力するクラス"""
    
    # 定数
    HEADER_ROWS = 6
    CATEGORY_COLUMNS = 3
    DATA_START_ROW = 7
    
    # カラー設定
    COLOR_INCREASED = '51CF66'  # 緑 (増加)
    COLOR_DECREASED = 'FF6B6B'  # 赤 (減少)
    COLOR_UNCHANGED = 'FFFFFF'  # 白 (変更なし)
    
    def __init__(self, output_dir: Path, sheet_name: str):
        """
        初期化
        
        Args:
            output_dir: 出力ディレクトリ
            sheet_name: シート名
        """
        self.output_dir = output_dir
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
    
    def write_diff_result(
        self,
        header_df: pd.DataFrame,
        category_df: pd.DataFrame,
        month_diffs: List[MonthlySalesMonthDiff],
        source_data: Dict
    ) -> Path:
        """
        差分結果を Excel ファイルとして出力
        
        Args:
            header_df: ヘッダー DataFrame (行1-6)
            category_df: カテゴリ DataFrame (列A-C)
            month_diffs: 月別差分のリスト
            source_data: 元データ (新ファイルから取得、grouping情報用)
            
        Returns:
            出力ファイルパス
        """
        # 出力ファイル名を生成
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{self.sheet_name}_差分_{timestamp}.xlsx"
        output_path = self.output_dir / output_filename
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = f"{self.sheet_name}_差分"
        
        # 1. ヘッダー部分を書き込み (行1-6, 全列)
        self._write_header(header_df, month_diffs)
        
        # 2. カテゴリ列を書き込み (列A-C, 全行)
        self._write_categories(category_df)
        
        # 3. 月別データを書き込み
        self._write_month_blocks(month_diffs)
        
        # 4. スタイルを適用
        self._apply_styles(month_diffs)
        
        # 5. 列幅を調整
        self._adjust_column_widths()
        
        # 6. フリーズペインを設定 (ヘッダー6行 + カテゴリ3列を固定)
        self.ws.freeze_panes = self.ws.cell(row=self.DATA_START_ROW, column=self.CATEGORY_COLUMNS + 1)
        
        # 保存
        self.wb.save(output_path)
        self.wb.close()
        
        return output_path
    
    def _write_header(self, header_df: pd.DataFrame, month_diffs: List[MonthlySalesMonthDiff]):
        """ヘッダー部分を書き込み"""
        # カテゴリ列のヘッダー (A-C列) を書き込み
        for row_idx in range(len(header_df)):
            for col_idx in range(self.CATEGORY_COLUMNS):
                cell_value = header_df.iloc[row_idx, col_idx]
                self.ws.cell(
                    row=row_idx + 1,
                    column=col_idx + 1,
                    value=cell_value
                )
        
        # 各月ブロックのヘッダーを書き込み (差分がある月のみ)
        for month_idx, month_diff in enumerate(month_diffs):
            source_start_col = month_diff.column_range[0]  # 元ファイルの列位置 (0-indexed)
            
            # 出力ファイルでの開始列を計算 (カテゴリ3列 + これまでの月ブロック)
            output_start_col = self.CATEGORY_COLUMNS + (month_idx * 4)
            
            # このブロックの4列分のヘッダーを書き込み
            for row_idx in range(len(header_df)):
                for col_offset in range(4):  # 4列/ブロック
                    source_col = source_start_col + col_offset
                    
                    # 行5 (index 4) の最初の列には月名を書き込み
                    if row_idx == 4 and col_offset == 0:
                        cell_value = month_diff.month_name
                    elif source_col < len(header_df.columns):
                        cell_value = header_df.iloc[row_idx, source_col]
                    else:
                        cell_value = None
                    
                    if cell_value is not None:
                        self.ws.cell(
                            row=row_idx + 1,
                            column=output_start_col + col_offset + 1,  # 1-indexed
                            value=cell_value
                        )
    
    def _write_categories(self, category_df: pd.DataFrame):
        """カテゴリ列を書き込み"""
        for row_idx in range(len(category_df)):
            for col_idx in range(self.CATEGORY_COLUMNS):
                cell_value = category_df.iloc[row_idx, col_idx]
                cell = self.ws.cell(
                    row=row_idx + 1,
                    column=col_idx + 1,
                    value=cell_value
                )
                # Note: カテゴリ列(A-C)のスタイルは _apply_styles() で設定
    
    def _write_month_blocks(self, month_diffs: List[MonthlySalesMonthDiff]):
        """月別データブロックを書き込み"""
        for month_idx, month_diff in enumerate(month_diffs):
            # 出力ファイルでの開始列 (カテゴリ3列 + これまでの月ブロック)
            output_start_col = self.CATEGORY_COLUMNS + (month_idx * 4)
            
            # セル差分を行・列でグループ化
            for cell_diff in month_diff.cell_diffs:
                row = self.DATA_START_ROW + cell_diff.row_index  # 7行目から開始
                
                # 列インデックスを計算 (売上=0, 外部原価=1, 内部原価=2, 営業利益=3)
                column_names = ['売上', '外部原価', '内部原価', '営業利益']
                col_offset = column_names.index(cell_diff.column_name)
                col = output_start_col + col_offset + 1  # 1-indexed
                
                # セルに値を書き込み
                formatted_text = cell_diff.get_formatted_text()
                cell = self.ws.cell(row=row, column=col, value=formatted_text)
                
                # 背景色を設定
                if cell_diff.change_type == 'increased':
                    cell.fill = PatternFill(start_color=self.COLOR_INCREASED, end_color=self.COLOR_INCREASED, fill_type='solid')
                elif cell_diff.change_type == 'decreased':
                    cell.fill = PatternFill(start_color=self.COLOR_DECREASED, end_color=self.COLOR_DECREASED, fill_type='solid')
                else:
                    # unchanged の場合は数値のみ表示
                    if cell_diff.new_value is not None:
                        cell.value = cell_diff.new_value
                        cell.number_format = '#,##0'
    
    def _apply_styles(self, month_diffs: List[MonthlySalesMonthDiff]):
        """スタイルを適用"""
        # フォント設定
        default_font = Font(name='Segoe UI', size=10)
        
        # 罫線設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 全セルに適用 (デフォルト)
        for row in self.ws.iter_rows():
            for cell in row:
                cell.font = default_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # ヘッダー行 (1-6) を太字 + 灰色背景
        for row_idx in range(1, self.HEADER_ROWS + 1):
            for cell in self.ws[row_idx]:
                cell.font = Font(name='Segoe UI', size=10, bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # カテゴリB列 (col 2) に値がある行 → 行全体を灰色背景 + 太字
        for row_idx in range(self.DATA_START_ROW, self.ws.max_row + 1):
            col_b_cell = self.ws.cell(row=row_idx, column=2)  # B列
            
            # B列に実際のデータがある場合のみ (None・NaN・空文字・空白を除く)
            has_data = False
            if col_b_cell.value is not None:
                # pandas の NaN をチェック
                if isinstance(col_b_cell.value, float):
                    import math
                    has_data = not math.isnan(col_b_cell.value)
                else:
                    has_data = str(col_b_cell.value).strip() != ''
            
            if has_data:
                # 行全体にスタイルを適用
                for cell in self.ws[row_idx]:
                    # 既存の背景色を保持 (増減の色がある場合)
                    if cell.fill.start_color.rgb in [None, '00000000', 'FFFFFFFF']:
                        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    # 太字にする
                    cell.font = Font(name='Segoe UI', size=10, bold=True)
        
        # 数値列を右寄せに
        for month_idx, month_diff in enumerate(month_diffs):
            # 出力ファイルでの列範囲
            output_start_col = self.CATEGORY_COLUMNS + (month_idx * 4) + 1  # 1-indexed
            output_end_col = output_start_col + 3  # 4列ブロック
            
            for col in range(output_start_col, output_end_col + 1):
                for row_idx in range(self.DATA_START_ROW, self.ws.max_row + 1):
                    cell = self.ws.cell(row=row_idx, column=col)
                    # 既存のフォント・背景色を保持したまま右寄せに
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    def _adjust_column_widths(self):
        """列幅を調整"""
        # カテゴリ列の幅
        self.ws.column_dimensions['A'].width = 20
        self.ws.column_dimensions['B'].width = 25
        self.ws.column_dimensions['C'].width = 30
        
        # データ列の幅 (自動調整)
        for col_idx in range(self.CATEGORY_COLUMNS + 1, self.ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            
            # 列内の最大文字数を取得
            max_length = 0
            for cell in self.ws[col_letter]:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # 幅を設定 (最小12, 最大50)
            adjusted_width = min(max(max_length + 2, 12), 50)
            self.ws.column_dimensions[col_letter].width = adjusted_width
    
    def _copy_grouping(self, source_ws: Worksheet):
        """
        元のワークシートからグループ化情報をコピー (実装は複雑なため後回し)
        
        Args:
            source_ws: 元のワークシート
        """
        # TODO: openpyxl でグループ化情報をコピーする実装
        # 現時点では未実装 (手動でグループ化する必要あり)
        pass
